"""Downloadable report generation services for MoneyLeak AI."""

from __future__ import annotations

import csv
import io
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from services.analytics_utils import (
    display_merchant_name,
    enum_to_string,
    get_category,
    get_field,
    parse_date,
    to_float,
    valid_transactions,
)
from services.burn_rate_analyzer import analyze_burn_rate
from services.dashboard_service import get_category_breakdown, get_needs_wants_waste, get_summary, get_top_merchants
from services.duplicate_detector import detect_duplicates
from services.leakage_detector import detect_small_spend_leakage
from services.merchant_analyzer import get_yearly_impact
from services.money_leak_score import calculate_score
from services.report_summary import generate_saving_priority
from services.smart_alerts import detect_bill_reminders, detect_refund_reversal_tracking
from services.monthly_explainer import explain_latest_month_change
from services.subscription_detector import detect_subscriptions

REPORT_LIMITATION_NOTICE = "This analysis is based on your uploaded statement period. It is for budgeting guidance only, not financial advice."
NO_SECTION_DATA = "No data available for this section"
CSV_COLUMNS = [
    "date",
    "merchant",
    "description",
    "amount",
    "type",
    "category",
    "confidence",
    "is_subscription",
    "is_duplicate",
    "is_refund",
    "flags",
]


class ReportGenerationError(ValueError):
    """Raised when a report cannot be generated from safe input."""


def first_name(value: Any) -> str:
    """Return only the first name from a user-like object or fallback label."""
    name = str(get_field(value, "full_name", "User") or "User").strip()
    if not name:
        return "User"
    return name.split()[0]


def iso_date(value: Any) -> str:
    """Return an ISO date string or an empty string for missing dates."""
    parsed = parse_date(value)
    return parsed.isoformat() if parsed else ""


def display_date(value: Any) -> str:
    """Return a readable date string or a neutral fallback."""
    parsed = parse_date(value)
    if parsed is None:
        return "Not available"
    return parsed.strftime("%d %b %Y")


def money(value: Any) -> str:
    """Format a numeric value as Indian rupee text for reports."""
    amount = round(to_float(value), 2)
    sign = "-" if amount < 0 else ""
    absolute = abs(amount)
    rupees = int(absolute)
    paise = int(round((absolute - rupees) * 100))
    formatted = f"{rupees:,}"
    parts = formatted.split(",")
    if len(parts) > 1:
        formatted = parts[0] + "," + ",".join(parts[1:])
        if len(parts[-1]) == 3:
            last_three = str(rupees)[-3:]
            leading = str(rupees)[:-3]
            if leading:
                leading_groups = []
                while len(leading) > 2:
                    leading_groups.insert(0, leading[-2:])
                    leading = leading[:-2]
                if leading:
                    leading_groups.insert(0, leading)
                formatted = ",".join(leading_groups + [last_three])
    return f"{sign}₹{formatted}.{paise:02d}"


def percent(value: Any) -> str:
    """Format a percentage value for report tables."""
    return f"{to_float(value):.2f}%"


def statement_period_from_inputs(statement: Any, transactions: list[Any]) -> dict[str, str | None]:
    """Resolve report period from statement metadata and transaction dates."""
    start_value = get_field(statement, "statement_period_start", None)
    end_value = get_field(statement, "statement_period_end", None)
    if start_value is not None or end_value is not None:
        return {"start": iso_date(start_value) or None, "end": iso_date(end_value) or None}
    dates = [parsed for parsed in (parse_date(get_field(transaction, "transaction_date", None)) for transaction in valid_transactions(transactions)) if parsed]
    if not dates:
        return {"start": None, "end": None}
    return {"start": min(dates).isoformat(), "end": max(dates).isoformat()}


def period_label(period: dict[str, str | None]) -> str:
    """Return a display label for a statement period dictionary."""
    start = display_date(period.get("start")) if period.get("start") else "Not available"
    end = display_date(period.get("end")) if period.get("end") else "Not available"
    if start == "Not available" and end == "Not available":
        return "Not available"
    return f"{start} to {end}"


def transaction_type_label(transaction: Any) -> str:
    """Return debit or credit label for a transaction."""
    raw_value = enum_to_string(get_field(transaction, "transaction_type", ""), "").lower()
    if raw_value in {"debit", "credit"}:
        return raw_value
    amount = to_float(get_field(transaction, "amount", 0.0))
    return "credit" if amount > 0 else "debit"


def transaction_flags(transaction: Any) -> str:
    """Return compact comma-separated transaction flags."""
    flags = []
    if bool(get_field(transaction, "is_subscription", False)):
        flags.append("subscription")
    if bool(get_field(transaction, "is_duplicate", False)):
        flags.append("duplicate")
    if bool(get_field(transaction, "is_small_spend", False)):
        flags.append("small_spend")
    if bool(get_field(transaction, "is_refund", False)) or bool(get_field(transaction, "is_cashback", False)):
        flags.append("refund")
    if bool(get_field(transaction, "needs_review", False)):
        flags.append("needs_review")
    if bool(get_field(transaction, "is_late_night", False)):
        flags.append("late_night")
    return ", ".join(flags)


def serialize_transaction_for_export(transaction: Any) -> dict[str, Any]:
    """Serialize one transaction using the CSV export column contract."""
    return {
        "date": iso_date(get_field(transaction, "transaction_date", None)),
        "merchant": display_merchant_name(get_field(transaction, "merchant", None)),
        "description": str(get_field(transaction, "description", "") or ""),
        "amount": round(to_float(get_field(transaction, "amount", 0.0)), 2),
        "type": transaction_type_label(transaction),
        "category": get_category(transaction),
        "confidence": round(to_float(get_field(transaction, "category_confidence", 0.0)), 4),
        "is_subscription": bool(get_field(transaction, "is_subscription", False)),
        "is_duplicate": bool(get_field(transaction, "is_duplicate", False)),
        "is_refund": bool(get_field(transaction, "is_refund", False) or get_field(transaction, "is_cashback", False)),
        "flags": transaction_flags(transaction),
    }


def build_report_data(transactions: list[Any], user: Any = None, statement: Any = None, budget: Any = None, current_balance: float | None = None) -> dict[str, Any]:
    """Build all analytics sections required by PDF and Excel reports."""
    safe_transactions = valid_transactions(transactions)
    subscriptions_result = detect_subscriptions(safe_transactions)
    duplicates_result = detect_duplicates(safe_transactions)
    subscriptions = subscriptions_result.get("data", [])
    duplicates = duplicates_result.get("data", [])
    score_result = calculate_score(safe_transactions, subscriptions, duplicates, budget)
    summary_result = get_summary(safe_transactions, budget)
    summary_data = summary_result.get("data", {}) or {}
    score_data = score_result.get("data", {}) or {}
    if isinstance(summary_data, dict):
        summary_data["money_leak_score"] = score_data
    leakage_result = detect_small_spend_leakage(safe_transactions)
    category_result = get_category_breakdown(safe_transactions)
    merchant_result = get_top_merchants(safe_transactions, limit=5)
    need_result = get_needs_wants_waste(safe_transactions)
    priority_result = generate_saving_priority(safe_transactions, subscriptions, duplicates)
    burn_result = analyze_burn_rate(safe_transactions, current_balance=current_balance)
    yearly_result = get_yearly_impact(safe_transactions)
    bill_result = detect_bill_reminders(safe_transactions)
    refund_result = detect_refund_reversal_tracking(safe_transactions)
    change_result = explain_latest_month_change(safe_transactions)
    period = statement_period_from_inputs(statement, safe_transactions)
    warnings = []
    for result in (subscriptions_result, duplicates_result, score_result, summary_result, leakage_result, category_result, merchant_result, need_result, priority_result, burn_result, yearly_result, bill_result, refund_result, change_result):
        warnings.extend(result.get("warnings", []))
    return {
        "user_first_name": first_name(user),
        "statement_period": period,
        "summary": summary_data,
        "category_breakdown": category_result.get("data", []),
        "top_merchants": merchant_result.get("data", []),
        "leakage": leakage_result.get("data", {}),
        "subscriptions": subscriptions,
        "duplicates": duplicates,
        "needs_wants_waste": need_result.get("data", {}),
        "saving_priority": priority_result.get("data", []),
        "burn_rate": burn_result.get("data", {}),
        "yearly_impact": yearly_result.get("data", []),
        "bill_reminders": bill_result.get("data", {}),
        "refund_tracking": refund_result.get("data", {}),
        "monthly_change_explanation": change_result.get("data", {}),
        "warnings": list(dict.fromkeys(str(warning) for warning in warnings if warning)),
    }


def safe_table_rows(rows: list[list[Any]], empty_message: str = NO_SECTION_DATA) -> list[list[Any]]:
    """Return report table rows or a single fallback row."""
    if len(rows) <= 1:
        width = len(rows[0]) if rows else 1
        return [[empty_message] + [""] * max(0, width - 1)]
    return rows


def pdf_table(rows: list[list[Any]], column_widths: list[float] | None = None) -> Table:
    """Create a consistently styled ReportLab table."""
    display_rows = [[str(cell if cell is not None else "") for cell in row] for row in safe_table_rows(rows)]
    table = Table(display_rows, colWidths=column_widths, repeatRows=1 if len(display_rows) > 1 else 0)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0F172A")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#CBD5E1")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
                ("LEFTPADDING", (0, 0), (-1, -1), 5),
                ("RIGHTPADDING", (0, 0), (-1, -1), 5),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]
        )
    )
    return table


def add_heading(story: list[Any], text: str, styles: dict[str, ParagraphStyle]) -> None:
    """Append a section heading to a PDF story."""
    story.append(Spacer(1, 0.14 * inch))
    story.append(Paragraph(text, styles["Heading2"]))
    story.append(Spacer(1, 0.06 * inch))


def no_data_paragraph(styles: dict[str, ParagraphStyle]) -> Paragraph:
    """Return a PDF paragraph stating that a section has no data."""
    return Paragraph(NO_SECTION_DATA, styles["BodyText"])


def build_pdf_report(transactions: list[Any], user: Any = None, statement: Any = None, budget: Any = None, current_balance: float | None = None) -> bytes:
    """Generate the monthly MoneyLeak AI PDF report as bytes."""
    buffer = io.BytesIO()
    document = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=36)
    base_styles = getSampleStyleSheet()
    base_styles.add(ParagraphStyle(name="Small", parent=base_styles["BodyText"], fontSize=8, leading=10))
    data = build_report_data(transactions, user=user, statement=statement, budget=budget, current_balance=current_balance)
    story: list[Any] = []

    story.append(Paragraph("MoneyLeak AI", base_styles["Title"]))
    story.append(Paragraph("Monthly Money Report", base_styles["Heading1"]))
    story.append(Paragraph(f"User: {data['user_first_name']}", base_styles["BodyText"]))
    story.append(Paragraph(f"Statement period: {period_label(data['statement_period'])}", base_styles["BodyText"]))

    summary = data.get("summary", {}) if isinstance(data.get("summary"), dict) else {}
    score = summary.get("money_leak_score", {}) if isinstance(summary.get("money_leak_score"), dict) else {}
    add_heading(story, "1. Executive Summary", base_styles)
    story.append(
        pdf_table(
            [
                ["Metric", "Value"],
                ["Total Spent", money(summary.get("total_spent"))],
                ["Total Received", money(summary.get("total_received"))],
                ["Net Balance", money(summary.get("net_balance_change"))],
                ["Money Leak Score", f"{score.get('score', 0)} ({score.get('severity', 'Healthy')})"],
                ["Diagnosis", score.get("diagnosis") or "No diagnosis available"],
            ],
            [2.0 * inch, 4.4 * inch],
        )
    )

    add_heading(story, "2. Spending Categories", base_styles)
    category_rows = [["Category", "Amount", "Percentage", "Type"]]
    for item in data.get("category_breakdown", []) or []:
        category_rows.append([item.get("category"), money(item.get("total_amount")), percent(item.get("percentage_of_total_spend")), item.get("need_want_waste_type")])
    story.append(pdf_table(category_rows, [2.3 * inch, 1.4 * inch, 1.2 * inch, 1.2 * inch]))

    add_heading(story, "3. Top 5 Merchants", base_styles)
    merchant_rows = [["Merchant", "Category", "Spent", "Transactions"]]
    for item in data.get("top_merchants", [])[:5]:
        merchant_rows.append([item.get("merchant"), item.get("category"), money(item.get("total_spent")), item.get("transaction_count")])
    story.append(pdf_table(merchant_rows, [2.3 * inch, 1.7 * inch, 1.2 * inch, 1.0 * inch]))

    add_heading(story, "4. Money Leaks", base_styles)
    leakage = data.get("leakage", {}) if isinstance(data.get("leakage"), dict) else {}
    subscriptions = data.get("subscriptions", []) or []
    duplicates = data.get("duplicates", []) or []
    subscription_yearly = sum(to_float(item.get("yearly_cost")) for item in subscriptions)
    duplicate_amount = sum(to_float(item.get("amount")) for item in duplicates)
    story.append(
        pdf_table(
            [
                ["Leak Type", "Value"],
                ["Small spend total", money(leakage.get("total_leakage"))],
                ["Subscriptions", f"{len(subscriptions)} detected, {money(subscription_yearly)} yearly"],
                ["Duplicate payments", f"{len(duplicates)} detected, {money(duplicate_amount)} total"],
            ],
            [2.4 * inch, 4.0 * inch],
        )
    )

    add_heading(story, "5. Subscriptions", base_styles)
    subscription_rows = [["Merchant", "Frequency", "Monthly", "Yearly", "Priority"]]
    for item in subscriptions:
        subscription_rows.append([item.get("merchant"), item.get("frequency"), money(item.get("monthly_cost")), money(item.get("yearly_cost")), item.get("cancellation_priority")])
    story.append(pdf_table(subscription_rows, [1.8 * inch, 1.1 * inch, 1.1 * inch, 1.1 * inch, 1.0 * inch]))

    add_heading(story, "6. Duplicate Payments", base_styles)
    duplicate_rows = [["Merchant", "Amount", "Date", "Confidence"]]
    for item in duplicates:
        duplicate_rows.append([item.get("merchant"), money(item.get("amount")), item.get("duplicate_date"), item.get("confidence_score")])
    story.append(pdf_table(duplicate_rows, [2.3 * inch, 1.4 * inch, 1.4 * inch, 1.2 * inch]))

    add_heading(story, "7. Needs vs Wants vs Waste", base_styles)
    need_data = data.get("needs_wants_waste", {}) if isinstance(data.get("needs_wants_waste"), dict) else {}
    story.append(
        pdf_table(
            [
                ["Type", "Total", "Percentage"],
                ["Needs", money(need_data.get("needs_total")), percent(need_data.get("needs_pct"))],
                ["Wants", money(need_data.get("wants_total")), percent(need_data.get("wants_pct"))],
                ["Waste", money(need_data.get("waste_total")), percent(need_data.get("waste_pct"))],
                ["Savings", money(need_data.get("savings_total")), percent(need_data.get("savings_pct"))],
            ],
            [2.1 * inch, 1.7 * inch, 1.5 * inch],
        )
    )

    story.append(PageBreak())
    add_heading(story, "8. Saving Priority List", base_styles)
    priorities = data.get("saving_priority", []) or []
    if not priorities:
        story.append(no_data_paragraph(base_styles))
    for item in priorities:
        story.append(Paragraph(f"{item.get('rank')}. {item.get('action')} — monthly saving {money(item.get('possible_monthly_saving'))}", base_styles["BodyText"]))
        story.append(Spacer(1, 0.04 * inch))

    add_heading(story, "9. Month-End Analysis", base_styles)
    burn = data.get("burn_rate", {}) if isinstance(data.get("burn_rate"), dict) else {}
    story.append(
        pdf_table(
            [
                ["Metric", "Value"],
                ["Daily burn rate", money(burn.get("daily_burn_rate"))],
                ["Daily safe limit", money(burn.get("daily_safe_limit")) if burn.get("daily_safe_limit") is not None else "Not available"],
                ["Days until empty", burn.get("days_until_empty") if burn.get("days_until_empty") is not None else "Not available"],
                ["Will survive month", burn.get("will_survive_month") if burn.get("will_survive_month") is not None else "Not available"],
            ],
            [2.4 * inch, 3.0 * inch],
        )
    )

    add_heading(story, "10. Yearly Impact", base_styles)
    yearly_rows = [["Category", "Period Amount", "Annualized Amount", "Type"]]
    for item in data.get("yearly_impact", []) or []:
        yearly_rows.append([item.get("category"), money(item.get("period_amount")), money(item.get("annualized_amount")), item.get("need_want_waste_type")])
    story.append(pdf_table(yearly_rows, [2.1 * inch, 1.4 * inch, 1.6 * inch, 1.1 * inch]))

    add_heading(story, "11. Bill and Renewal Reminders", base_styles)
    bill_data = data.get("bill_reminders", {}) if isinstance(data.get("bill_reminders"), dict) else {}
    reminder_rows = [["Merchant", "Category", "Amount", "Predicted Due", "Priority"]]
    for item in (bill_data.get("reminders", []) or [])[:12]:
        reminder_rows.append([item.get("merchant"), item.get("category"), money(item.get("amount")), item.get("predicted_due_date"), item.get("priority")])
    story.append(pdf_table(reminder_rows, [1.8 * inch, 1.3 * inch, 1.1 * inch, 1.4 * inch, 1.0 * inch]))

    add_heading(story, "12. Refund and Reversal Tracking", base_styles)
    refund_data = data.get("refund_tracking", {}) if isinstance(data.get("refund_tracking"), dict) else {}
    review_rows = [["Merchant", "Amount", "Date", "Status", "Reason"]]
    for item in (refund_data.get("review_items", []) or [])[:12]:
        review_rows.append([item.get("merchant"), money(item.get("amount")), item.get("transaction_date"), item.get("status"), item.get("reason")])
    story.append(pdf_table(review_rows, [1.5 * inch, 0.9 * inch, 1.0 * inch, 0.9 * inch, 2.1 * inch]))

    add_heading(story, "13. What Changed This Month", base_styles)
    change_data = data.get("monthly_change_explanation", {}) if isinstance(data.get("monthly_change_explanation"), dict) else {}
    story.append(Paragraph(str(change_data.get("headline") or "At least two months are needed to explain month-over-month changes."), base_styles["BodyText"]))
    driver_rows = [["Driver", "Explanation"]]
    for index, item in enumerate(change_data.get("drivers", []) or [], start=1):
        driver_rows.append([index, item])
    story.append(pdf_table(driver_rows, [0.7 * inch, 5.4 * inch]))

    add_heading(story, "14. Limitations Notice", base_styles)
    story.append(Paragraph(REPORT_LIMITATION_NOTICE, base_styles["Small"]))
    document.build(story)
    return buffer.getvalue()


def build_csv_export(transactions: list[Any]) -> bytes:
    """Generate CSV transaction export bytes using the required columns."""
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=CSV_COLUMNS, extrasaction="ignore")
    writer.writeheader()
    for transaction in valid_transactions(transactions):
        writer.writerow(serialize_transaction_for_export(transaction))
    return output.getvalue().encode("utf-8-sig")


def append_key_value_rows(sheet: Any, rows: list[tuple[str, Any]]) -> None:
    """Append key-value rows to an Excel worksheet."""
    for key, value in rows:
        sheet.append([key, value])


def write_table(sheet: Any, headers: list[str], rows: list[list[Any]]) -> None:
    """Write headers and rows to an Excel worksheet."""
    sheet.append(headers)
    if not rows:
        sheet.append([NO_SECTION_DATA] + [""] * max(0, len(headers) - 1))
        return
    for row in rows:
        sheet.append(row)


def style_worksheet(sheet: Any) -> None:
    """Apply navy headers, alternating rows, and auto-width to an Excel worksheet."""
    navy_fill = PatternFill("solid", fgColor="0F172A")
    white_font = Font(color="FFFFFF", bold=True)
    alternate_fill = PatternFill("solid", fgColor="F8FAFC")
    for cell in sheet[1]:
        cell.fill = navy_fill
        cell.font = white_font
    for row_index, row in enumerate(sheet.iter_rows(min_row=2), start=2):
        if row_index % 2 == 0:
            for cell in row:
                cell.fill = alternate_fill
    for column_cells in sheet.columns:
        column_letter = get_column_letter(column_cells[0].column)
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 48)


def build_excel_report(transactions: list[Any], user: Any = None, statement: Any = None, budget: Any = None, current_balance: float | None = None) -> bytes:
    """Generate a multi-sheet Excel report as bytes."""
    safe_transactions = valid_transactions(transactions)
    data = build_report_data(safe_transactions, user=user, statement=statement, budget=budget, current_balance=current_balance)
    workbook = Workbook()

    summary_sheet = workbook.active
    summary_sheet.title = "Summary"
    summary = data.get("summary", {}) if isinstance(data.get("summary"), dict) else {}
    score = summary.get("money_leak_score", {}) if isinstance(summary.get("money_leak_score"), dict) else {}
    summary_sheet.append(["Metric", "Value"])
    append_key_value_rows(
        summary_sheet,
        [
            ("User", data.get("user_first_name") or "User"),
            ("Statement Period", period_label(data.get("statement_period", {}))),
            ("Total Spent", summary.get("total_spent", 0.0)),
            ("Total Received", summary.get("total_received", 0.0)),
            ("Net Balance", summary.get("net_balance_change", 0.0)),
            ("Money Leak Score", score.get("score", 0.0)),
            ("Severity", score.get("severity", "Healthy")),
            ("Diagnosis", score.get("diagnosis", "No diagnosis available")),
            ("Limitations", REPORT_LIMITATION_NOTICE),
        ],
    )

    transaction_sheet = workbook.create_sheet("Transactions")
    write_table(transaction_sheet, CSV_COLUMNS, [[row[column] for column in CSV_COLUMNS] for row in [serialize_transaction_for_export(transaction) for transaction in safe_transactions]])

    category_sheet = workbook.create_sheet("Category Breakdown")
    write_table(
        category_sheet,
        ["category", "amount", "percentage", "type", "transactions", "average"],
        [
            [item.get("category"), item.get("total_amount"), item.get("percentage_of_total_spend"), item.get("need_want_waste_type"), item.get("transaction_count"), item.get("average_transaction_amount")]
            for item in data.get("category_breakdown", [])
        ],
    )

    subscription_sheet = workbook.create_sheet("Subscriptions")
    write_table(
        subscription_sheet,
        ["merchant", "frequency", "average_amount", "monthly_cost", "yearly_cost", "priority", "next_predicted_date"],
        [
            [item.get("merchant"), item.get("frequency"), item.get("average_amount"), item.get("monthly_cost"), item.get("yearly_cost"), item.get("cancellation_priority"), item.get("next_predicted_date")]
            for item in data.get("subscriptions", [])
        ],
    )

    leaks_sheet = workbook.create_sheet("Money Leaks")
    leakage = data.get("leakage", {}) if isinstance(data.get("leakage"), dict) else {}
    leaks_sheet.append(["Leak Type", "Count", "Amount", "Details"])
    leaks_sheet.append(["Small spends under 100", (leakage.get("bucket_under_100") or {}).get("count", 0), (leakage.get("bucket_under_100") or {}).get("total", 0.0), "amount < 100"])
    leaks_sheet.append(["Small spends 100 to 200", (leakage.get("bucket_100_to_200") or {}).get("count", 0), (leakage.get("bucket_100_to_200") or {}).get("total", 0.0), "100 <= amount < 200"])
    leaks_sheet.append(["Small spends 200 to 500", (leakage.get("bucket_200_to_500") or {}).get("count", 0), (leakage.get("bucket_200_to_500") or {}).get("total", 0.0), "200 <= amount < 500"])
    for item in data.get("duplicates", []) or []:
        leaks_sheet.append(["Duplicate payment", 1, item.get("amount", 0.0), f"{item.get('merchant')} on {item.get('duplicate_date')} confidence {item.get('confidence_score')}"])

    recommendation_sheet = workbook.create_sheet("Saving Recommendations")
    write_table(
        recommendation_sheet,
        ["rank", "target", "reason", "monthly_saving", "yearly_saving", "difficulty", "action"],
        [
            [item.get("rank"), item.get("target"), item.get("reason"), item.get("possible_monthly_saving"), item.get("possible_yearly_saving"), item.get("difficulty"), item.get("action")]
            for item in data.get("saving_priority", [])
        ],
    )

    bill_sheet = workbook.create_sheet("Bill Reminders")
    bill_data = data.get("bill_reminders", {}) if isinstance(data.get("bill_reminders"), dict) else {}
    write_table(
        bill_sheet,
        ["merchant", "category", "amount", "last_paid_date", "predicted_due_date", "days_until_due", "status", "priority", "reason"],
        [
            [item.get("merchant"), item.get("category"), item.get("amount"), item.get("last_paid_date"), item.get("predicted_due_date"), item.get("days_until_due"), item.get("status"), item.get("priority"), item.get("reason")]
            for item in bill_data.get("reminders", []) or []
        ],
    )

    refund_sheet = workbook.create_sheet("Refund Tracking")
    refund_data = data.get("refund_tracking", {}) if isinstance(data.get("refund_tracking"), dict) else {}
    write_table(
        refund_sheet,
        ["merchant", "amount", "transaction_date", "status", "matched_refund_date", "matched_refund_amount", "reason"],
        [
            [item.get("merchant"), item.get("amount"), item.get("transaction_date"), item.get("status"), item.get("matched_refund_date"), item.get("matched_refund_amount"), item.get("reason")]
            for item in refund_data.get("review_items", []) or []
        ],
    )

    change_sheet = workbook.create_sheet("Month Changes")
    change_data = data.get("monthly_change_explanation", {}) if isinstance(data.get("monthly_change_explanation"), dict) else {}
    change_sheet.append(["Metric", "Value"])
    append_key_value_rows(
        change_sheet,
        [
            ("From Month", change_data.get("from_label")),
            ("To Month", change_data.get("to_label")),
            ("Spending Change", change_data.get("spending_change")),
            ("Income Change", change_data.get("income_change")),
            ("Headline", change_data.get("headline")),
        ],
    )
    change_sheet.append([])
    write_table(change_sheet, ["driver"], [[item] for item in change_data.get("drivers", []) or []])

    for worksheet in workbook.worksheets:
        style_worksheet(worksheet)
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
