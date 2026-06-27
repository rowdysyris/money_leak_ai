"""Downloadable report generation tests for MoneyLeak AI."""

from __future__ import annotations

import csv
import io
from datetime import date
from decimal import Decimal
from types import SimpleNamespace
from uuid import uuid4

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from main import app
from services.report_generator import CSV_COLUMNS, build_csv_export, build_excel_report, build_pdf_report


def sample_user() -> SimpleNamespace:
    """Return a user-like object with extra name data that must not leak into reports."""
    return SimpleNamespace(id=uuid4(), full_name="Kedar Sensitive Surname", email="private@example.com")


def sample_statement(user_id: object | None = None) -> SimpleNamespace:
    """Return a statement-like object for report tests."""
    return SimpleNamespace(
        id=uuid4(),
        user_id=user_id or uuid4(),
        statement_period_start=date(2024, 1, 1),
        statement_period_end=date(2024, 1, 31),
        original_filename="statement.csv",
    )


def tx(
    merchant: str,
    amount: str,
    transaction_type: str,
    category: str,
    transaction_date: date = date(2024, 1, 15),
    confidence: float = 0.95,
    description: str | None = None,
    is_subscription: bool = False,
    is_duplicate: bool = False,
    is_refund: bool = False,
    need_want_waste_type: str = "want",
) -> SimpleNamespace:
    """Build a transaction-like object for report tests."""
    return SimpleNamespace(
        id=uuid4(),
        user_id=uuid4(),
        statement_id=uuid4(),
        transaction_date=transaction_date,
        transaction_time=None,
        merchant=merchant,
        description=description or merchant,
        amount=Decimal(amount),
        transaction_type=transaction_type,
        category=category,
        category_confidence=confidence,
        is_subscription=is_subscription,
        is_duplicate=is_duplicate,
        is_small_spend=abs(Decimal(amount)) < Decimal("500") and transaction_type == "debit",
        is_refund=is_refund,
        is_cashback=False,
        needs_review=False,
        is_late_night=False,
        need_want_waste_type=need_want_waste_type,
    )


def report_transactions() -> list[SimpleNamespace]:
    """Return a compact but varied transaction fixture for reports."""
    return [
        tx("Salary", "30000.00", "credit", "Transfers", date(2024, 1, 1), need_want_waste_type="unknown"),
        tx("Swiggy", "450.00", "debit", "Food & Dining", date(2024, 1, 5), need_want_waste_type="want"),
        tx("Netflix", "649.00", "debit", "Subscriptions", date(2024, 1, 10), is_subscription=True, need_want_waste_type="want"),
        tx("Netflix", "649.00", "debit", "Subscriptions", date(2024, 2, 9), is_subscription=True, need_want_waste_type="want"),
        tx("Bank Fee", "118.00", "debit", "Bank Charges & Fees", date(2024, 1, 12), need_want_waste_type="waste"),
        tx("Dominos", "899.00", "debit", "Food & Dining", date(2024, 1, 16), is_duplicate=True, need_want_waste_type="want"),
        tx("Dominos", "899.00", "debit", "Food & Dining", date(2024, 1, 16), is_duplicate=True, need_want_waste_type="want"),
    ]


def test_pdf_generates_successfully() -> None:
    """PDF report generation returns a valid PDF byte stream."""
    content = build_pdf_report(report_transactions(), user=sample_user(), statement=sample_statement())
    assert content.startswith(b"%PDF")
    assert len(content) > 1000


def test_pdf_with_no_transactions_graceful() -> None:
    """PDF report generation handles an empty transaction list without crashing."""
    content = build_pdf_report([], user=sample_user(), statement=sample_statement())
    assert content.startswith(b"%PDF")
    assert len(content) > 1000


def test_csv_exports_correct_columns() -> None:
    """CSV export contains the required exact column set."""
    content = build_csv_export(report_transactions())
    decoded = content.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(decoded))
    assert reader.fieldnames == CSV_COLUMNS
    rows = list(reader)
    assert rows[0]["date"] == "2024-01-01"
    assert "flags" in rows[0]


def test_excel_has_all_sheets() -> None:
    """Excel report contains all required worksheets."""
    content = build_excel_report(report_transactions(), user=sample_user(), statement=sample_statement())
    workbook = load_workbook(io.BytesIO(content))
    assert workbook.sheetnames == [
        "Summary",
        "Transactions",
        "Category Breakdown",
        "Subscriptions",
        "Money Leaks",
        "Saving Recommendations",
        "Bill Reminders",
        "Refund Tracking",
        "Month Changes",
    ]


def test_report_endpoints_require_auth() -> None:
    """Report endpoints require Bearer authentication."""
    client = TestClient(app)
    for endpoint in ("/api/reports/download/pdf", "/api/reports/download/csv", "/api/reports/download/excel"):
        response = client.get(endpoint)
        assert response.status_code == 401
        body = response.json()
        assert body["success"] is False
        assert body["error"]["code"] == "NOT_AUTHENTICATED"


def test_report_hides_sensitive_fields() -> None:
    """Exports avoid account numbers, bank names, email addresses, and full user names."""
    user = sample_user()
    statement = sample_statement(user.id)
    pdf_content = build_pdf_report(report_transactions(), user=user, statement=statement)
    csv_content = build_csv_export(report_transactions())
    excel_content = build_excel_report(report_transactions(), user=user, statement=statement)
    combined = pdf_content + csv_content + excel_content
    assert b"private@example.com" not in combined
    assert b"Sensitive Surname" not in combined
    assert b"account_number" not in combined.lower()
    assert b"bank_name" not in combined.lower()
